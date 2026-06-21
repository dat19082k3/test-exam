import sys
import os
import asyncio
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Add root project dir to path so we can import src modules easily
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "../..")))

from playwright.async_api import async_playwright, Page
from src.bot.consumer import get_5_star_books

async def search_wikipedia(page: Page, title: str) -> str:
    """Search for a title on Wikipedia.
    
    1. Navigate to https://en.wikipedia.org/
    2. Wait for search input (explicit wait)
    3. Type title into search bar
    4. Submit search
    5. Wait for results page
    6. Return URL of the first result page, or "No result"
    """
    try:
        if title == "Unknown" or not title:
            return "No result"
            
        # 1. Navigate
        await page.goto("https://en.wikipedia.org/")
        
        # 2. Wait for search input
        search_input = page.locator('input[name="search"]').first
        await search_input.wait_for(state="visible", timeout=10000)
        
        # 3. Type title
        await search_input.fill(title)
        
        # 4. Submit
        await search_input.press("Enter")
        
        # 5. Wait for results page to load
        # Wait for either the article text or search results to appear
        await page.wait_for_selector('#mw-content-text', timeout=15000)
        
        # 6. Check if we landed on a search results page instead of a direct article
        if "Special:Search" in page.url or "Special%3ASearch" in page.url or "search=" in page.url.lower():
            first_result = page.locator('.mw-search-result-heading a').first
            if await first_result.count() > 0:
                await first_result.click()
                await page.wait_for_selector('#mw-content-text', timeout=15000)
                return page.url
            else:
                return "No result"
                
        # If we didn't land on Special:Search, Wikipedia redirected us to the exact article match
        return page.url
        
    except Exception as e:
        print(f"  [ERROR] Failed to search '{title}': {e}")
        return "No result"

async def search_all_books(books: list[dict]) -> list[dict]:
    """Search Wikipedia for all 5-star books.
    Add 'wikipedia_url' field to each book dict.
    """
    if not books:
        return books
        
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        page = await browser.new_page()
        
        for idx, book in enumerate(books, 1):
            title = book.get("title", "")
            print(f"[{idx}/{len(books)}] Searching Wikipedia for book title: {title}...")
            
            url = await search_wikipedia(page, title)
            book["wikipedia_url"] = url
            print(f"  -> [OK] {url}")
                
        await browser.close()
        
    return books

def generate_report(books: list[dict], output_path: str = "output/rpa_report.xlsx") -> None:
    """Generate an Excel report for the 5-star books using openpyxl."""
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "5-Star Books Report"
    
    headers = ["Title", "Price", "Publisher Country", "Wikipedia URL"]
    ws.append(headers)
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
                         
    for col_num in range(1, 5):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        
    link_font = Font(color="0563C1", underline="single")
    
    for book in books:
        url = book.get("wikipedia_url", "")
        row = [
            book.get("title", ""),
            book.get("price", ""),
            book.get("country", ""),
            url
        ]
        ws.append(row)
        
        row_num = ws.max_row
        for col_num in range(1, 5):
            ws.cell(row=row_num, column=col_num).border = thin_border
            
        if url and url.startswith("http"):
            url_cell = ws.cell(row=row_num, column=4)
            url_cell.hyperlink = url
            url_cell.font = link_font
            
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 60)
        
    wb.save(output_path)
    print(f"\n[OK] Excel report successfully generated at: {output_path}")

async def run_wiki_bot():
    print("Step 1: Fetching 5-star books from API...")
    books = get_5_star_books()
    print(f"Found {len(books)} books.")
    
    print("\nStep 2: Automating Wikipedia via Playwright (Async)...")
    enriched_books = await search_all_books(books)
    
    print("\nStep 3: Generating Excel Report...")
    generate_report(enriched_books)
    
    return enriched_books

if __name__ == "__main__":
    enriched_books = asyncio.run(run_wiki_bot())
    print("\n--- Final Enriched Data (Preview) ---")
    for b in enriched_books[:3]:
        print(f"\nTitle: {b['title']}")
        print(f"Country: {b.get('country')}")
        print(f"Wikipedia URL: {b.get('wikipedia_url')}")
